from tkinter import *
from PIL import ImageTk, Image
import openpyxl
from openpyxl import load_workbook
import lxml
import csv
import pandas as pd

pd.set_option('display.max_columns', 13)


root = Tk()
root.title("Skill Matrix ")
photo1 = PhotoImage(file="amazonsmile.gif")
Label(root, image=photo1, bg="black").grid(row=0, column=0)
book = openpyxl.load_workbook("Skill Matrix.xlsx")
sheet = book.active

def selected(value):
	statement = Label(root, text="Choose a function")
	statement.grid(row=0,column=0)

def add_associate():
	top = Toplevel()
	top.title("Add associate")

	def clear():
		Login_entry.delete(0, END)
		Pick_text.delete(0, END)
		Pack_sin_text.delete(0, END)
		pack_mult_text.delete(0, END)
		Lift_op_entry.delete(0, END)
		cart_run_entry.delete(0, END)
		FUD_PG_entry.delete(0, END)
		Reactive_PG_entry.delete(0, END)
		instructor_entry.delete(0, END)
		prob_solve_entry.delete(0, END)
		Pick_text.delete(0, END)
		Pack_sin_text.delete(0, END)
		pack_mult_text.delete(0, END)
		receive_text.delete(0, END)
		stow_entry.delete(0, END)
		ICQA_entry.delete(0, END)
	
	def submit():
			newAA = [[Login_entry.get(), Lift_op_entry.get(), cart_run_entry.get(), FUD_PG_entry.get(), Reactive_PG_entry.get(), instructor_entry.get(), prob_solve_entry.get(), 
						 Pick_text.get(), Pack_sin_text.get(), pack_mult_text.get(), receive_text.get(), stow_entry.get(), ICQA_entry.get(), tkvar.get()]]
			for entry in newAA:
				sheet.append(entry)
				book.save('Skill Matrix.xlsx')
				clear()



	Login = Label(top, text="Login:", font="MsSerif 12 ")
	Login.grid(row=1, column=1, sticky=W)
	Login_entry = Entry(top, width=15)
	Login_entry.grid(row=1, column=2,)

	lift_op = Label(top, text="Lift OP:", font="MsSerif 12")
	lift_op.grid(row=2, column=1, sticky=W)
	Lift_op_entry = Entry(top, width=15)
	Lift_op_entry.grid(row=2,column=2)

	cart_run = Label(top, text="Cart Runner:", font="MsSerif 12")
	cart_run.grid(row=3, column=1, sticky=W)
	cart_run_entry = Entry(top, width=15)
	cart_run_entry.grid(row=3,column=2)

	FUD_PG = Label(top, text="FUD PG:", font="MsSerif 12")
	FUD_PG.grid(row=4, column=1,sticky=W)
	FUD_PG_entry = Entry(top, width=15)
	FUD_PG_entry.grid(row=4,column=2)

	Reactive_PG = Label(top, text="Reactive PG:", font="MsSerif 12 ")
	Reactive_PG.grid(row=5, column=1,sticky=W)
	Reactive_PG_entry = Entry(top, width=15)
	Reactive_PG_entry.grid(row=5, column=2)

	instructor = Label(top, text="Instructor:", font="MsSerif 12 ")
	instructor.grid(row=6, column=1, sticky=W)
	instructor_entry = Entry(top, width=15)
	instructor_entry.grid(row=6, column=2)

	prob_solve = Label(top, text="Problem Solve:", font="MsSerif 12 ")
	prob_solve.grid(row=7, column=1,sticky=W)
	prob_solve_entry = Entry(top, width=15)
	prob_solve_entry.grid(row=7, column=2)

	Pick = Label(top, text="Pick Trained:", font="MsSerif 12 ")
	Pick.grid(row=8, column=1,sticky=W)
	Pick_text = Entry(top, width =15)
	Pick_text.grid(row=8, column=2)

	Pack_sin = Label(top, text="Pack Singles:", font="MsSerif 12 ")
	Pack_sin.grid(row=9, column=1,sticky=W)
	Pack_sin_text = Entry(top, width=15)
	Pack_sin_text.grid(row=9, column=2)

	Pack_mult = Label(top, text="Pack Multis:", font="MsSerif 12 ")
	Pack_mult.grid(row=10, column=1,sticky=W)
	pack_mult_text = Entry(top, width=15)
	pack_mult_text.grid(row=10, column=2)

	receive = Label(top, text="Receive:", font="MsSerif 12 ")
	receive.grid(row=11, column=1,sticky=W)
	receive_text = Entry(top, width=15)
	receive_text.grid(row=11, column=2)

	stow = Label(top, text="Stow:", font="MsSerif 12 ")
	stow.grid(row=12, column=1,sticky=W)
	stow_entry = Entry(top, width=15)
	stow_entry.grid(row=12, column=2)

	ICQA = Label(top, text="ICQA:", font="MsSerif 12 ")
	ICQA.grid(row=13, column=1,sticky=W)
	ICQA_entry = Entry(top, width=15)
	ICQA_entry.grid(row=13, column=2)

	#Shift dropdown
	shift = Label(top, text="Shift Pattern:", font="MsSerif 12")
	shift.grid(row=14, column=1,sticky=W)
	shifts = ["Sun-Weds", "Mon-Thurs", "Tues-Fri", "Weds-Sat", "Thurs-Sun", "Fri-Mon", "Sat-Tues"]
	tkvar = StringVar(top)
	tkvar.set("Sun-Weds")
	shift_menu = OptionMenu(top, tkvar, *shifts)
	shift_menu.grid(row=14, column=2)




	#Buttons
	clearbutton = Button(top, text="Clear", command=clear, padx=5, pady=5)
	clearbutton.grid(row=15, column = 1)
	submit_button = Button(top, text="Submit", command=submit, padx=5, pady=5)
	submit_button.grid(row=15, column=2)

		

def delete_associate():
	global photo3

	def delete_ass():
		for index, row in enumerate(sheet.iter_rows()):
			for cell in row:
				if cell.value == del_search_entry.get():
					sheet.delete_rows(index+1,1)
					book.save('Skill Matrix.xlsx')

	top = Toplevel()
	top.title("Delete Associate")
	delLabel = Label(top, text="You have selected to delete an associate").grid(row=0,column=0)

	photo3 = PhotoImage(file="delete.gif")
	Label(top, image=photo3, bg="black").grid(row=0, column=0)
	del_search = Label(top, text="Login to delete:", font="MsSerif 12 ")
	del_search.grid(row=1, column=0)
	del_search_entry = Entry(top, width=15)
	del_search_entry.grid(row=2, column=0)
	del_button = Button(top, text="Delete",command=delete_ass, font="MsSerif 12")
	del_button.grid(row=3,column=0)


	
def search_associate_skills():
	global photo2
	top = Toplevel()
	top.title("Search Associate Skills")

	def search():
		excel_file = 'Skill Matrix.xlsx'
		skills = pd.read_excel(excel_file, index_col = "Login")
		received = ([search_entry.get()])
		first = skills.loc[received]
		top = Toplevel()
		top.title("Associate Skills")
		frame = LabelFrame(top, text="Associate Skills", padx=5, pady=5)
		frame.grid(row=1,column=1, padx=175,pady=35)
		second_ass_label = Label(top, text="Skill Set").grid(row=0,column=0)
		third_label = Label(top, text=first.to_string()).grid(row=1,column=0)
		print(first.to_string())
		


	photo2 = PhotoImage(file="mag_search.gif")
	Label(top, image=photo2, bg="black").grid(row=0, column=0)
	serLabel = Label(top, text="Enter Associate to Search").grid(row=12,column=0)
	search_entry = Entry(top, width=15)
	search_entry.grid(row=13, column=0)
	ser_button = Button(top, text="Submit Search", command=search, padx=5, pady=5)
	ser_button.grid(row=14, column=0)


add_ass = Button(root, text="Add Associate", command=add_associate).grid(row=1,column=0)
del_ass = Button(root, text="Delete Entry", command=delete_associate).grid(row=2,column=0)
search_ass =Button(root, text="Search Associate Skills", command=search_associate_skills).grid(row=3,column=0)





root.mainloop()